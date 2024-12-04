import asyncio
import re
from openpyxl import load_workbook
from datetime import datetime
import signal
import aiohttp
import json
from typing import Optional, Dict, Any
from scr.api_updaters.update_ozon import update_prices_ozon
from scr.api_updaters.update_wb import update_prices_wb
from scr.api_updaters.update_ym import update_price_ym
from scr.api_updaters.update_mm import update_prices_mm
from scr.logger import logger
from scr.data_fetcher import get_sheet_data, save_to_database
from scr.data_updater import update_prices ,update_and_merge_dataframes
from scr.data_writer import write_sheet_data
from scr.get_data.get_ozon_data import get_products_report,update_dataframe_ozon,sort_by_status_async
import pandas as pd
from scr.get_data.get_wb_data.get_first_wb_datas import get_wb_data
from scr.get_data.get_wb_data.get_wb_stocks import get_stocks
from scr.get_data.get_wb_data.wb_update_functions import update_stocks_data,update_dataframe_wb,clean_numeric_column
from scr.get_data.get_wb_data.images_wb import add_image_formulas



# Флаг для корректного завершения программы
is_running = True
DEBUG = True


class MarketplaceConfig:
    def __init__(self,
                 user_id: str,
                 sample_spreadsheet_id: str,
                 update_interval_minutes: int,
                 api_ozon: Optional[str] = None,
                 client_id_ozon: Optional[str] = None,
                 ozon_range: Optional[str] = None,
                 api_yandex_market: Optional[str] = None,
                 business_id_yandex_market: Optional[str] = None,
                 yandex_market_range: Optional[str] = None,
                 api_wildberries: Optional[str] = None,
                 wildberries_range: Optional[str] = None,
                 api_megamarket: Optional[str] = None,
                 megamarket_range: Optional[str] = None,
                 market_name: Optional[str] = None,
                 user_email: Optional[str] = None,
                 phone_number: Optional[str] = None):
        self.user_id = user_id
        self.sample_spreadsheet_id = sample_spreadsheet_id
        self.update_interval_minutes = int(update_interval_minutes)
        self.api_ozon = api_ozon
        self.client_id_ozon = client_id_ozon
        self.ozon_range = ozon_range
        self.api_yandex_market = api_yandex_market
        self.business_id_yandex_market = business_id_yandex_market
        self.yandex_market_range = yandex_market_range
        self.api_wildberries = api_wildberries
        self.wildberries_range = wildberries_range
        self.api_megamarket = api_megamarket
        self.megamarket_range = megamarket_range
        self.market_name = market_name
        self.user_email = user_email
        self.phone_number = phone_number
    def get_user_info(self) -> str:
        """Возвращает информацию о пользователе для логов"""
        return f"[ID: {self.user_id}, Email: {self.user_email}, Тел: {self.phone_number}]"

    def has_ozon_config(self) -> bool:
        return all([self.api_ozon, self.client_id_ozon, self.ozon_range])

    def has_yandex_market_config(self) -> bool:
        return all([self.api_yandex_market, self.business_id_yandex_market, self.yandex_market_range])

    def has_wildberries_config(self) -> bool:
        return all([self.api_wildberries, self.wildberries_range])

    def has_megamarket_config(self) -> bool:
        return all([self.api_megamarket, self.megamarket_range])


async def process_ozon_data(session: aiohttp.ClientSession, config: MarketplaceConfig) -> Dict[str, Any]:
    """Обработка данных Ozon"""
    ozon_logger = logger.bind(marketplace="Ozon")

    # Валидация входных данных
    if not all([config.sample_spreadsheet_id, config.market_name, config.ozon_range,
                config.client_id_ozon, config.api_ozon]):
        raise ValueError("Отсутствуют необходимые параметры конфигурации")

    try:
        # Инициализация параметров
        db_config = {
            'spreadsheet_id': config.sample_spreadsheet_id,
            'safe_market_name': re.sub(r'[^\w\-_]', '_', config.market_name),
            'range_name': config.market_name,
            'sheet_range': config.ozon_range,
            'client_id': str(config.client_id_ozon),
            'api_key': config.api_ozon,
            'safe_user_name': re.sub(r'[^\w\-_]', '_', config.user_id)
        }

        COLUMNS_FULL = {
            'id_col': 'id',
            'product_id_col': 'product_id',
            'price_col': 't_price',
            'old_price_col': 'price',
            'old_disc_in_base_col': 'price_old',
            'old_disc_manual_col': 'old_price',
            'min_price_base': 'min_price_old',
            'min_price': 'min_price',
            'prim_col': 'prim'
        }

        SQLITE_DB_NAME = f"databases/{db_config['safe_user_name']}_data_{db_config['safe_market_name']}.db"

        # 1. Получение данных из Google Sheets
        df_from_sheet = None
        try:
            ozon_logger.info(f"Получение данных из Google Sheets для {db_config['range_name']}")
            df_from_sheet = await get_sheet_data(db_config['spreadsheet_id'], db_config['sheet_range'])
        except Exception as e:
            ozon_logger.error(
                "Не удалось получить данные из Google Sheets",
                extra={
                    'user_id': config.user_id,
                    'email': config.user_email,
                    'error': str(e)
                }
            )

            # 2. Получение данных из Ozon
        report_df = pd.DataFrame()
        try:
            ozon_logger.info(f'Запрос данных из Ozon для пользователя {config.user_id}')
            report_df = await get_products_report(
                client_id=db_config['client_id'],
                api_key=db_config['api_key'],
                marketname=config.market_name,
                username=config.user_id,
            )
        except Exception as e:
            ozon_logger.error(
                "Критическая ошибка при запросе данных Ozon",
                extra={
                    'user_id': config.user_id,
                    'market_name': config.market_name,
                    'range': config.ozon_range,
                    'error': str(e)
                }
            )

            # 3. Обработка данных
        if df_from_sheet is not None and not report_df.empty:
            try:
                # Объединение данных
                df_united = await update_dataframe_ozon(df_from_sheet, report_df, config.user_id, config.market_name)
                df = await sort_by_status_async(df_united)

                # Сохранение в базу данных
                try:
                    await save_to_database(
                        df,
                        SQLITE_DB_NAME,
                        f'product_data_ozon_{db_config["range_name"]}',
                        primary_key_cols=['product_id']
                    )
                except Exception as e:
                    ozon_logger.error(f"Ошибка сохранения в базу данных: {str(e)}")

                    # Обновление цен
                updated_df, price_changed_df = await update_prices(
                    df=df,
                    columns_dict=COLUMNS_FULL,
                    marketplace='Ozone',
                    username=config.user_id,
                    sqlite_db_name=SQLITE_DB_NAME
                )

                # Запись обновленных данных
                await write_sheet_data(
                    updated_df,
                    db_config['spreadsheet_id'],
                    db_config['sheet_range'].replace('1', '3')
                )

                # Обновление цен через API
                if not price_changed_df.empty:
                    flag = await update_prices_ozon(
                        df=price_changed_df,
                        new_price_col="t_price",
                        base_old_price_col='price_old',
                        old_price_col="old_price",
                        product_id_col="product_id",
                        offer_id_col='offer_id',
                        min_price_col="min_price",
                        client_id=db_config['client_id'],
                        api_key=db_config['api_key'],
                        debug=DEBUG
                    )

                    if not flag:
                        df_from_error = await update_and_merge_dataframes(
                            updated_df,
                            price_changed_df,
                            'offer_id'
                        )
                        await write_sheet_data(
                            df_from_error,
                            db_config['spreadsheet_id'],
                            db_config['sheet_range'].replace('1', '3')
                        )

                return {
                    'status': 'success',
                    'marketplace': 'Ozon',
                    'rows_processed': len(df),
                    'rows_updated': len(price_changed_df)
                }

            except Exception as e:
                ozon_logger.error(
                    "Ошибка при обработке данных",
                    extra={
                        'user_id': config.user_id,
                        'market_name': config.market_name,
                        'error': str(e)
                    }
                )
                raise

        elif not report_df.empty:
            # Если нет данных в Google Sheets, но есть данные из Ozon
            ozon_logger.warning(
                f"Таблица Google пуста для клиента {config.user_id}, записываю данные из личного кабинета"
            )
            # Создаем копию DataFrame
            df_to_write = report_df.copy()

            # Сортируем с сохранением описательной строки в начале
            df_to_write = await sort_by_status_async(df_to_write)

            # Получаем названия колонок
            column_names = pd.DataFrame([df_to_write.columns.tolist()], columns=df_to_write.columns)

            # Сначала добавляем названия колонок, затем данные
            df_to_write = pd.concat([column_names, df_to_write], axis=0, ignore_index=True)



            await write_sheet_data(
                df_to_write,
                db_config['spreadsheet_id'],
                db_config['sheet_range']
            )
        return {
                'status': 'success',
                'marketplace': 'Ozon',
                'rows_processed': len(report_df),
                'rows_updated': 0
            }

        return {
            'status': 'error',
            'marketplace': 'Ozon',
            'error': 'Нет данных для обработки',
            'details': {
                'has_sheet_data': df_from_sheet is not None,
                'has_ozon_data': not report_df.empty
            }
        }

    except Exception as e:
        return {
            'status': 'error',
            'marketplace': 'Ozon',
            'error': str(e),
            'details': {
                'user_id': config.user_id,
                'market_name': config.market_name,
                'range': config.ozon_range
            }
        }


async def process_yandex_market_data(session: aiohttp.ClientSession, config: MarketplaceConfig) -> Dict[str, Any]:
    """Обработка данных Яндекс.Маркет"""
    ym_logger = logger.bind(marketplace="YandexMarket")

    # Валидация входных данных
    if not all([config.sample_spreadsheet_id, config.market_name, config.yandex_market_range,
                config.api_yandex_market, config.business_id_yandex_market]):
        raise ValueError("Отсутствуют необходимые параметры конфигурации Яндекс.Маркет")

    try:
        # Инициализация параметров
        db_config = {
            'spreadsheet_id': config.sample_spreadsheet_id,
            'safe_market_name': re.sub(r'[^\w\-_]', '_', config.market_name),
            'range_name': config.market_name,
            'sheet_range': config.yandex_market_range,
            'api_key': config.api_yandex_market,
            'business_id': config.business_id_yandex_market,
            'safe_user_name': re.sub(r'[^\w\-_]', '_', config.user_id)
        }
        COLUMNS_FULL = {
            'id_col': 'id',
            'product_id_col': 'offer_id',
            'price_col': 't_price',
            'old_price_col': 'price',
            'old_disc_in_base_col': 'price_old',
            'old_disc_manual_col': 'discount_base',
            'prim_col': 'prim'
        }
        SQLITE_DB_NAME = f"databases/{db_config['safe_user_name']}_data_{db_config['safe_market_name']}.db"


        try:
            # Получение данныхym_logger.info(f"Получение данных из Google Sheets для {db_config['range_name']}")
            df = await get_sheet_data(db_config['spreadsheet_id'], db_config['sheet_range'])

            if df is None or df.empty:
                raise ValueError(f"Не удалось получить данные из Google Sheets для {db_config['range_name']}")
        except Exception as e:
            ym_logger.error(f"Ошибка получения данных из Google Sheets для пользователя {config.user_id} "
                            f"с email {config.user_email}", error=str(e))
            raise

        try:
            # Сохранение исходных данных
            ym_logger.info(f"Сохранение данных в базу для {db_config['range_name']}")
            await save_to_database(
                df,
                SQLITE_DB_NAME,
                f"product_data_ym_{db_config['safe_market_name']}",
                primary_key_cols=['offer_id']
            )
        except Exception as e:
            ym_logger.error(f"Ошибка сохранения в базу данных {SQLITE_DB_NAME}", error=str(e))
            raise

            # Обновление цен
        ym_logger.info(f"Обновление цен для {db_config['range_name']}")
        updated_df, price_changed_df = await update_prices(
            df=df,
            columns_dict=COLUMNS_FULL,
            marketplace='Yandex Market',
            username=config.user_id,
            sqlite_db_name=SQLITE_DB_NAME
        )

        # Запись обновленных данных
        ym_logger.info(f"Запись обновленных данных в Google Sheets для {db_config['range_name']}")
        await write_sheet_data(
            updated_df,
            db_config['spreadsheet_id'],
            db_config['sheet_range'].replace('1', '3')
        )

        # Обновление цен через API если есть изменения
        if not price_changed_df.empty:
            ym_logger.warning(
                f"Обновление цен через API Яндекс.Маркет для {db_config['range_name']}",
                rows_to_update=len(price_changed_df)
            )
            flag = await update_price_ym(
                df=price_changed_df,
                access_token=db_config['api_key'],
                campaign_id=db_config['business_id'],
                offer_id_col="offer_id",
                disc_old_col="price_old",
                new_price_col="t_price",
                discount_base_col="discount_base",
                debug=DEBUG
            )
            if flag == False :
                df_from_error = await update_and_merge_dataframes(updated_df,
                                                            price_changed_df,'offer_id')
                await write_sheet_data(
                    df_from_error,
                    db_config['spreadsheet_id'],
                    db_config['sheet_range'].replace('1', '3'))

        ym_logger.info(
            f"Обработка завершена успешно для {db_config['range_name']}",
            rows_processed=len(df),
            rows_updated=len(price_changed_df)
        )

        return {
            'status': 'success',
            'marketplace': 'YandexMarket',
            'rows_processed': len(df),
            'rows_updated': len(price_changed_df)
        }

    except Exception as e:
        error_details = {
            'user_id': config.user_id,
            'market_name': config.market_name,
            'range': config.yandex_market_range,
            'business_id': config.business_id_yandex_market,
            'error': str(e)
        }
        ym_logger.error("Критическая ошибка при обновлении данных Яндекс.Маркет", **error_details)


        return {
            'status': 'error',
            'marketplace': 'YandexMarket',
            'error': str(e),
            'details': error_details
        }


async def process_wildberries_data(session: aiohttp.ClientSession, config: MarketplaceConfig) -> Dict[str, Any]:
    """Обработка данных Wildberries"""
    wb_logger = logger.bind(marketplace="Wildberries")

    # Валидация входных данных
    if not all([config.sample_spreadsheet_id, config.market_name,
                config.wildberries_range, config.api_wildberries]):
        raise ValueError("Отсутствуют необходимые параметры конфигурации Wildberries")

    try:
        # Инициализация параметров
        db_config = {
            'spreadsheet_id': config.sample_spreadsheet_id,
            'safe_market_name': re.sub(r'[^\w\-_]', '_', config.market_name),
            'range_name': config.market_name,
            'sheet_range': config.wildberries_range,
            'api_key': config.api_wildberries ,
            'safe_user_name': re.sub(r'[^\w\-_]', '_', config.user_id)
        }
        COLUMNS_FULL = {
            'id_col': 'id',
            'product_id_col': 'nmID',
            'price_col': 't_price',
            'old_price_col': 'price',
            'old_disc_in_base_col': 'disc_old',
            'old_disc_manual_col': 'discount',
            'prim_col': 'prim'
        }
        SQLITE_DB_NAME = f"databases/{db_config['safe_user_name']}_data_{db_config['safe_market_name']}.db"

        try:
            # Получение данных гугл
            wb_logger.info(f"Получение данных из Google Sheets для {db_config['range_name']}")
            df_from_sheets = None
            df_from_sheets = await get_sheet_data(db_config['spreadsheet_id'], db_config['sheet_range'])

            try:
                df_from_sheets = await clean_numeric_column(old_df=df_from_sheets,
                                                               column_name='nmID',
                                                               username=config.user_id,
                                                               marketname=config.market_name,
                                                               logger=logger)
            except Exception as e:
                wb_logger.error(f"Не удалось удаление пустых и невалидных строк для пользователя {config.user_id}", error=str(e))



        except Exception as e:
            df_from_sheets = None
            wb_logger.error(f"Ошибка получения данных из Google Sheets для пользователя {config.user_id} "
                            f"с email {config.user_email}", error=str(e))



        # Получение данных ВБ
        try:
            first_df = await get_wb_data(api_key = db_config['api_key'],
                                         limit = 200,
                                         username=config.user_id,
                                         marketname=config.market_name)
        except Exception as e:
            wb_logger.error(
                "Критическая ошибка при запросе данных WB",
                extra={
                    'user_id': config.user_id,
                    'market_name': config.market_name,
                    'range': config.wildberries_range,
                    'error': str(e)
                }
            )
            first_df = None
        # Получение информации об остатках товаров
        try:
            stocks_df = await get_stocks(api_key=db_config['api_key'],
                                         marketname=config.market_name,
                                         username=config.user_id)
        except Exception as e:
            wb_logger.error(
                "Критическая ошибка при запросе данных WB",
                extra={
                    'user_id': config.user_id,
                    'market_name': config.market_name,
                    'range': config.wildberries_range,
                    'error': str(e)
                }
            )
            stocks_df = None
        if df_from_sheets is not None and first_df is not None:
            g_data = True
            upddated_df_without_stocks = await update_dataframe_wb(df_from_sheets,
                                                                  first_df,
                                                                  config.user_id,
                                                                  config.market_name)
            if stocks_df is not None:
                upddated_df_final = await update_stocks_data(upddated_df_without_stocks,
                                                             stocks_df,
                                                             logger)
            else:
                upddated_df_final = upddated_df_without_stocks.copy()
                wb_logger.warning(f"Ошибка обновления остатков для WB пользователь {config.user_id},диапазон{config.wildberries_range}")

        elif df_from_sheets is None and first_df is not None:
            g_data = False
            upddated_df_without_stocks = first_df.copy()

            if stocks_df is not None:
                upddated_df_final = await update_stocks_data(upddated_df_without_stocks,
                                                             stocks_df,
                                                             logger)
            else:
                upddated_df_final = upddated_df_without_stocks.copy()
                wb_logger.warning(f"Ошибка обновления остатков для WB пользователь {config.user_id},диапазон{config.wildberries_range}")
        else:
            upddated_df_final = None
            wb_logger.error(f"Не удалось получить ни Данные гугл таблиц,ни данные гугл")


        if upddated_df_final is not None:

            try:
                # Сохранение исходных данных
                wb_logger.info(f"Сохранение данных в базу для {db_config['range_name']}")
                await save_to_database(
                    upddated_df_final,
                    SQLITE_DB_NAME,
                    f"product_data_wb_{db_config['safe_market_name']}",
                    primary_key_cols=['nmID']
                )

            except Exception as e:
                wb_logger.error(f"Ошибка сохранения в базу данных {SQLITE_DB_NAME}", error=str(e))
                raise


        if upddated_df_final is not None and g_data == True:
            # Обновление цен
            wb_logger.info(f"Обновление цен для {db_config['range_name']}")
            updated_df, price_changed_df = await update_prices(
                df=upddated_df_final,
                columns_dict=COLUMNS_FULL,
                marketplace='Wildberries',
                username=config.user_id,
                sqlite_db_name=SQLITE_DB_NAME
            )
            updated_df = await add_image_formulas(updated_df)
            # Запись обновленных данных
            wb_logger.info(f"Запись обновленных данных в Google Sheets для {db_config['range_name']}")
            await write_sheet_data(
                updated_df,
                db_config['spreadsheet_id'],
                db_config['sheet_range'].replace('1', '3')
            )
            # Обновление цен через API если есть изменения
            if not price_changed_df.empty:
                wb_logger.warning(
                    f"Обновление цен через API Wildberries для {db_config['range_name']}",
                    rows_to_update=len(price_changed_df)
                )
                flag = await update_prices_wb(
                    df=price_changed_df,
                    nmID_col="nmID",
                    price_col="t_price",
                    discount_col="discount",
                    disc_old_col='disc_old',
                    api_key=db_config['api_key'],
                    debug=DEBUG
                )
                if flag == False :
                    df_from_error = await update_and_merge_dataframes(updated_df,
                                                                price_changed_df,'nmID')
                    await write_sheet_data(
                        df_from_error,
                        db_config['spreadsheet_id'],
                        db_config['sheet_range'].replace('1', '3')
                    )

            wb_logger.info(
                f"Обработка завершена успешно для {db_config['range_name']}",
                rows_processed=len(upddated_df_final),
                rows_updated=len(price_changed_df)
            )

            return {
                'status': 'success',
                'marketplace': 'Wildberries',
                'rows_processed': len(upddated_df_final),
                'rows_updated': len(price_changed_df)
            }

        elif upddated_df_final is not None and g_data == False:
            # Если нет данных в Google Sheets, но есть данные из WB
            wb_logger.warning(
                f"Таблица Google пуста для клиента {config.user_id}, записываю данные из личного кабинета"
            )
            # Создаем копию DataFrame
            df_to_write = upddated_df_final.copy()


            # Получаем названия колонок
            column_names = pd.DataFrame([df_to_write.columns.tolist()], columns=df_to_write.columns)



            df_to_write = await add_image_formulas(df_to_write)

            # Сначала добавляем названия колонок, затем данные
            df_to_write = pd.concat([column_names, df_to_write], axis=0, ignore_index=True)

            await write_sheet_data(
                df_to_write,
                db_config['spreadsheet_id'],
                db_config['sheet_range']
            )
        return {
            'status': 'success',
            'marketplace': 'WB',
            'rows_processed': len(upddated_df_final),
            'rows_updated': 0
        }


    except Exception as e:
        error_details = {
            'user_id': config.user_id,
            'market_name': config.market_name,
            'range': config.wildberries_range,
            'error': str(e)
        }
        wb_logger.error("Критическая ошибка при обновлении данных Wildberries", **error_details)


        return {
            'status': 'error',
            'marketplace': 'Wildberries',
            'error': str(e),
            'details': error_details
        }


async def process_megamarket_data(session: aiohttp.ClientSession, config: MarketplaceConfig) -> Dict[str, Any]:
    """Обработка данных Megamarket"""
    mm_logger = logger.bind(marketplace="Megamarket")

    # Валидация входных данных
    if not all([config.sample_spreadsheet_id, config.market_name,
                config.megamarket_range, config.api_megamarket]):
        raise ValueError("Отсутствуют необходимые параметры конфигурации Megamarket")

    try:
        # Инициализация параметров
        db_config = {
            'spreadsheet_id': config.sample_spreadsheet_id,
            'safe_market_name': re.sub(r'[^\w\-_]', '_', config.market_name),
            'range_name': config.market_name,
            'sheet_range': config.megamarket_range,
            'api_token': config.api_megamarket,
            'safe_user_name': re.sub(r'[^\w\-_]', '_', config.user_id)
        }
        COLUMNS_FULL = {
            'id_col': 'id',
            'product_id_col': 'seller_id',
            'price_col': 't_price',
            'old_price_col': 'price',
            'prim_col': 'prim'
        }
        SQLITE_DB_NAME = f"databases/{db_config['safe_user_name']}_data_{db_config['safe_market_name']}.db"

        try:
            # Получение данных
            mm_logger.info(f"Получение данных из Google Sheets для {db_config['range_name']}")
            df = await get_sheet_data(db_config['spreadsheet_id'], db_config['sheet_range'])

            if df is None or df.empty:
                raise ValueError(
                    f"Не удалось получить данные из Google Sheets для {db_config['range_name']}")
        except Exception as e:
            mm_logger.error(
                f"Ошибка получения данных из Google Sheets для пользователя {config.user_id} "f"с email {config.user_email}",
                error=str(e))
            raise

        try:
            # Сохранение исходных данных
            mm_logger.info(f"Сохранение данных в базу для {db_config['range_name']}")
            await save_to_database(
                df,
                SQLITE_DB_NAME,
                f"product_data_mm_{db_config['safe_market_name']}",
                primary_key_cols=['seller_id']
            )
        except Exception as e:
            mm_logger.error(f"Ошибка сохранения в базу данных {SQLITE_DB_NAME}", error=str(e))
            raise

            # Обновление цен
        mm_logger.info(f"Обновление цен для {db_config['range_name']}")
        updated_df, price_changed_df = await update_prices(
            df=df,
            columns_dict=COLUMNS_FULL,
            marketplace='MegaMarket',
            username=config.user_id,
            sqlite_db_name=SQLITE_DB_NAME
        )

        # Запись обновленных данных
        mm_logger.info(f"Запись обновленных данных в Google Sheets для {db_config['range_name']}")
        await write_sheet_data(
            updated_df,
            db_config['spreadsheet_id'],
            db_config['sheet_range'].replace('1', '3')
        )

        # Обновление цен через API если есть изменения
        if not price_changed_df.empty:
            mm_logger.warning(
                f"Обновление цен через API Megamarket для {db_config['range_name']}",
                rows_to_update=len(price_changed_df)
            )
            flag = await update_prices_mm(
                df=price_changed_df,
                token=db_config['api_token'],
                offer_id_col="seller_id",
                price_col = "t_price",
                debug=DEBUG
            )
            if flag == False :
                df_from_error = await update_and_merge_dataframes(updated_df,
                                                            price_changed_df,'seller_id')
                await write_sheet_data(
                    df_from_error,
                    db_config['spreadsheet_id'],
                    db_config['sheet_range'].replace('1', '3')
                )


        mm_logger.info(
            f"Обработка завершена успешно для {db_config['range_name']}",
            rows_processed=len(df),
            rows_updated=len(price_changed_df)
        )

        return {
            'status': 'success',
            'marketplace': 'Megamarket',
            'rows_processed': len(df),
            'rows_updated': len(price_changed_df)
        }

    except Exception as e:
        error_details = {
            'user_id': config.user_id,
            'market_name': config.market_name,
            'range': config.megamarket_range,
            'error': str(e)
        }
        mm_logger.error("Критическая ошибка при обновлении данных Megamarket", **error_details)


        return {
            'status': 'error',
            'marketplace': 'Megamarket',
            'error': str(e),
            'details': error_details
        }


def get_users_config_from_excel(filename: str) -> list:
    """Читает конфигурацию пользователей из Excel файла"""
    try:
        wb = load_workbook(filename=filename, read_only=True)
        sheet = wb.active
        # Получаем заголовки (ID пользователей)
        users = [cell.value for cell in sheet[1][1:] if cell.value]

        # Определяем индексы строк для каждого параметра
        param_indices = {}
        required_params = [
            'SAMPLE_SPREADSHEET_ID',
            'UPDATE_INTERVAL_MINUTES',
            'API_OZON',
            'CLIENT_ID_OZON',
            'OZON_RANGE',
            'API_YANDEX_MARKET',
            'BUSINESS_ID_YANDEX_MARKET',
            'YANDEX_MARKET_RANGE',
            'API_WILDBERRIES',
            'WILDBERRIES_RANGE',
            'API_MEGAMARKET',
            'MEGAMARKET_RANGE',
            'MARKET_NAME',
            'USER_EMAIL',
            'PHONE_NUMBER'
        ]

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            param_name = row[0].value
            if param_name in required_params:
                param_indices[param_name] = row_idx

        user_configs = []
        for user_id in users:
            user_column = None
            for idx, cell in enumerate(sheet[1]):
                if cell.value == user_id:
                    user_column = idx + 1
                    break

            if user_column:
                params = {}
                for param_name, row_idx in param_indices.items():
                    cell_value = sheet.cell(row=row_idx, column=user_column).value
                    params[param_name.lower()] = cell_value

                config = MarketplaceConfig(
                    user_id=user_id,
                    sample_spreadsheet_id=params.get('sample_spreadsheet_id'),
                    update_interval_minutes=params.get('update_interval_minutes', 5),
                    api_ozon=params.get('api_ozon'),
                    client_id_ozon=params.get('client_id_ozon'),
                    ozon_range=params.get('ozon_range'),
                    api_yandex_market=params.get('api_yandex_market'),
                    business_id_yandex_market=params.get('business_id_yandex_market'),
                    yandex_market_range=params.get('yandex_market_range'),
                    api_wildberries=params.get('api_wildberries'),
                    wildberries_range=params.get('wildberries_range'),
                    api_megamarket=params.get('api_megamarket'),
                    megamarket_range=params.get('megamarket_range'),
                    market_name=params.get('market_name'),
                    user_email=params.get('user_email'),
                    phone_number=params.get('phone_number')
                )
                user_configs.append(config)

        return user_configs

    except Exception as e:
        logger.error(f"❌ Ошибка при чтении конфигурации из Excel файла: {str(e)}")
        raise


async def process_marketplace_data(config: MarketplaceConfig):
    """Асинхронная функция обработки данных маркетплейсов"""
    user_info = config.get_user_info()
    try:
        async with aiohttp.ClientSession() as session:
            while is_running:
                start_time = datetime.now()
                logger.warning(f"🔄 Начало обработки данных для пользователя {user_info}")

                try:
                    results = []
                    # Обработка данных Ozon
                    if config.has_ozon_config():
                        logger.info(f"📦 Обработка данных OZON для пользователя {user_info}")
                        result = await process_ozon_data(session, config)
                        results.append(result)

                    # # Обработка данных Яндекс.Маркет
                    # if config.has_yandex_market_config():
                    #     logger.info(f"🎁 Обработка данных Яндекс.Маркет для пользователя {user_info}")
                    #     result = await process_yandex_market_data(session, config)
                    #     results.append(result)

                    # Обработка данных Wildberries
                    if config.has_wildberries_config():
                        logger.info(f"🛍️ Обработка данных Wildberries для пользователя {user_info}")
                        result = await process_wildberries_data(session, config)
                        results.append(result)
                    #
                    # # Обработка данных Megamarket
                    # if config.has_megamarket_config():
                    #     logger.info(f"🏪 Обработка данных Megamarket для пользователя {user_info}")
                    #     result = await process_megamarket_data(session, config)
                    #     results.append(result)

                    logger.info(f"✅ Завершена обработка данных для пользователя {user_info}")
                    logger.debug(f"Результаты обработки: {json.dumps(results, indent=2)}")

                except Exception as e:
                    logger.error(f"❌ Ошибка при обработке данных для пользователя {user_info}: {str(e)}")

                # Вычисляем время до следующего запуска
                processing_time = (datetime.now() - start_time).total_seconds()
                sleep_time = max(0, config.update_interval_minutes * 60 - processing_time)

                logger.info(f"💤 Пользователь {user_info} -ожидание {sleep_time:.1f} секунд до следующего обновления")
                await asyncio.sleep(sleep_time)

    except asyncio.CancelledError:
        logger.info(f"🛑 Задача для пользователя {user_info} была отменена")
    except Exception as e:
        logger.error(f"❌ Критическая ошибка в обработке данных для пользователя {user_info}: {str(e)}")


def handle_shutdown(signum, frame):
    """Обработчик сигналов завершения"""
    global is_running
    logger.info("🛑 Получен сигнал завершения. Начинаем корректное завершение всех задач...")
    is_running = False


async def main():
    """Основная функция"""
    try:
        # Получаем конфигурации всех пользователей
        user_configs = get_users_config_from_excel('config.xlsx')
        logger.info(f"🚀 Запуск обработки данных для {len(user_configs)} пользователей")

        # Создаем и запускаем задачи для всех пользователей
        tasks = []
        for config in user_configs:
            task = asyncio.create_task(
                process_marketplace_data(config),
                name=f"task_{config.user_id}"
            )
            tasks.append(task)
            logger.info(f"✨ Создана задача для пользователя {config.get_user_info()}")

        # Ждем завершения всех задач
        logger.info("⚡ Все задачи запущены. Нажмите Ctrl+C для остановки.")
        await asyncio.gather(*tasks, return_exceptions=True)

    except Exception as e:
        logger.error(f"❌Ошибка в главной функции: {str(e)}")
    finally:
        logger.info("🏁 Работа программы завершена")


if __name__ == "__main__":
    # Регистрируем обработчики сигналов для корректного завершения
    signal.signal(signal.SIGINT, handle_shutdown)
    signal.signal(signal.SIGTERM, handle_shutdown)

    # Запускаем асинхронное выполнение
    logger.info("🎯 Запуск программы обработки данных маркетплейсов")
    asyncio.run(main())